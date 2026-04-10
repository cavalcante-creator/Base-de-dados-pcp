import os
import re
from datetime import datetime
from io import BytesIO, StringIO

import pandas as pd
import pdfplumber
import pytz
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="PCP Produção", layout="wide")

fuso = pytz.timezone("America/Sao_Paulo")


def agora():
    return datetime.now(fuso)


def exportar_excel_formatado(df, nome_aba="Dados"):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nome_aba)
        ws = writer.book[nome_aba]

        ws.freeze_panes = "A2"

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if ws.max_row >= 2 and ws.max_column >= 1:
            ultima_coluna = get_column_letter(ws.max_column)
            nome_tabela = re.sub(r"\W+", "", nome_aba) or "Dados"
            tabela = Table(
                displayName=f"Tabela{nome_tabela[:20]}",
                ref=f"A1:{ultima_coluna}{ws.max_row}"
            )
            tabela.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(tabela)

        for idx, coluna in enumerate(df.columns, start=1):
            letra_coluna = get_column_letter(idx)
            valores = [str(coluna)]
            valores.extend("" if pd.isna(valor) else str(valor) for valor in df[coluna])
            ws.column_dimensions[letra_coluna].width = min(max(len(valor) for valor in valores) + 2, 40)

            if pd.api.types.is_numeric_dtype(df[coluna]):
                for cell in ws[letra_coluna][1:]:
                    cell.number_format = "#,##0.00"

        if "Status" in df.columns:
            idx_status = list(df.columns).index("Status") + 1
            for cell in ws[get_column_letter(idx_status)][1:]:
                valor = str(cell.value or "")
                if "FALTA" in valor:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
                elif "RISCO" in valor:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
                elif "OK" in valor:
                    cell.fill = PatternFill(fill_type="solid", fgColor="C6E0B4")

    output.seek(0)
    return output.getvalue()


def botao_downloads(df, nome_base, nome_aba):
    col_csv, col_excel = st.columns(2)

    with col_csv:
        st.download_button(
            f"Baixar {nome_base} CSV",
            df.to_csv(index=False).encode("utf-8"),
            file_name=f"{nome_base}.csv",
            mime="text/csv",
            key=f"csv_{nome_base}_{nome_aba}"
        )

    with col_excel:
        st.download_button(
            f"Baixar {nome_base} Excel",
            exportar_excel_formatado(df, nome_aba=nome_aba[:31]),
            file_name=f"{nome_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"excel_{nome_base}_{nome_aba}"
        )


def ir_para(pagina):
    st.query_params["pagina"] = pagina
    st.rerun()


def pagina_atual():
    pagina = st.query_params.get("pagina", "principal")
    if isinstance(pagina, list):
        return pagina[0]
    return pagina


def menu_lateral():
    st.sidebar.title("Navegação")

    if st.sidebar.button("Página Principal", use_container_width=True):
        ir_para("principal")

    if st.sidebar.button("Dashboard PCP", use_container_width=True):
        ir_para("dashboard")

    st.sidebar.markdown("---")
    st.sidebar.write("Endereços:")
    st.sidebar.code("?pagina=principal")
    st.sidebar.code("?pagina=dashboard")


def render_principal():
    st.title("PCP Produção")
    st.caption("Página principal do sistema")

    abas = st.tabs([
        "Saldo",
        "Perfil",
        "Ordens",
        "Previsão",
        "Base de Dados"
    ])

    with abas[0]:
        st.title("Saldo Produção")

        file = st.file_uploader("PDF Saldo", type=["pdf"])

        if file:
            with open("saldo_temp.pdf", "wb") as f:
                f.write(file.read())

            if st.button("Processar Saldo"):
                linhas = []

                with pdfplumber.open("saldo_temp.pdf") as pdf:
                    for p in pdf.pages:
                        texto = p.extract_text()
                        if texto:
                            linhas.extend(texto.split("\n"))

                dados = {}
                codigo_atual = None

                for linha in linhas:
                    linha = linha.strip()
                    if not linha:
                        continue

                    if "ALMOXARIFADO" in linha.upper():
                        codigo_match = None
                    else:
                        codigo_match = re.search(r"\b([A-Z]{1,3}\d{3,5})\b", linha)

                    if codigo_match:
                        codigo_atual = codigo_match.group(1)
                        if codigo_atual not in dados:
                            dados[codigo_atual] = {
                                "Codigo": codigo_atual,
                                "Saldo Total": 0,
                                "Saldo Almox 3": 0
                            }
                        continue

                    if "ALMOXARIFADO" in linha.upper() and codigo_atual:
                        nums = re.findall(r"[\d\.]+\,\d+", linha)
                        if nums:
                            valor = float(nums[-1].replace(".", "").replace(",", "."))
                            dados[codigo_atual]["Saldo Total"] += valor

                            if re.search(r"ALMOXARIFADO\s*:\s*3\b", linha.upper()):
                                dados[codigo_atual]["Saldo Almox 3"] += valor

                df = pd.DataFrame(dados.values())
                df["Data Processamento"] = agora().strftime("%d/%m/%Y")
                df["Hora Processamento"] = agora().strftime("%H:%M:%S")

                df.to_csv("saldo.csv", index=False)

                st.success("Saldo processado!")
                st.dataframe(df, use_container_width=True)

    with abas[1]:
        st.title("Perfil Produção")

        file = st.file_uploader("PDF Perfil", type=["pdf"])

        if file:
            with open("perfil_temp.pdf", "wb") as f:
                f.write(file.read())

            if st.button("Processar Perfil"):
                movimentacoes = []
                codigo_item = ""

                regex = re.compile(
                    r"(DD|DC|DP).*?(\d{2}/\d{2}/\d{4}).*?(-?[\d,.]+)"
                )

                with pdfplumber.open("perfil_temp.pdf") as pdf:
                    for p in pdf.pages:
                        texto = p.extract_text()
                        if texto:
                            for linha in texto.split("\n"):
                                item_match = re.search(r"Item:\s*(\S+)", linha)
                                if item_match:
                                    codigo_item = item_match.group(1)

                                mov = regex.search(linha)
                                if mov:
                                    movimentacoes.append({
                                        "Item": codigo_item,
                                        "Tipo": mov.group(1),
                                        "Data Fim": mov.group(2),
                                        "Quantidade": mov.group(3)
                                    })

                df = pd.DataFrame(movimentacoes)
                df["Data Processamento"] = agora().strftime("%d/%m/%Y")
                df["Hora Processamento"] = agora().strftime("%H:%M:%S")

                df.to_csv("perfil.csv", index=False)

                st.success("Perfil processado!")
                st.dataframe(df, use_container_width=True)

    with abas[2]:
        st.title("Ordens")

        file = st.file_uploader("CSV Ordens", type=["csv"])

        if file:
            conteudo = file.read().decode("utf-8", errors="ignore")
            df = pd.read_csv(StringIO(conteudo), sep=None, engine="python")

            df["Data Processamento"] = agora().strftime("%d/%m/%Y")
            df["Hora Processamento"] = agora().strftime("%H:%M:%S")

            df.to_csv("ordens.csv", index=False)

            st.success("Ordens carregadas!")
            st.dataframe(df, use_container_width=True)

    with abas[3]:
        st.title("Previsão")

        file = st.file_uploader("Excel Previsão", type=["xlsx"])

        if file:
            df_raw = pd.read_excel(file, header=None)

            linha_header = None
            for i in range(len(df_raw)):
                if "COD" in df_raw.iloc[i].astype(str).str.upper().values:
                    linha_header = i
                    break

            if linha_header is None:
                st.error("Não foi possível localizar a linha de cabeçalho.")
                st.stop()

            df = pd.read_excel(file, header=linha_header)
            df.columns = df.columns.astype(str).str.upper().str.strip()

            col_cod = [c for c in df.columns if "COD" in c][0]
            col_prod = [c for c in df.columns if "PROD" in c][0]

            df = df[[col_cod, col_prod]]
            df.columns = ["COD", "PRODUTO"]

            df["Data Processamento"] = agora().strftime("%d/%m/%Y")
            df["Hora Processamento"] = agora().strftime("%H:%M:%S")

            df.to_csv("previsao.csv", index=False)

            st.success("Previsão carregada!")
            st.dataframe(df, use_container_width=True)

    with abas[4]:
        st.title("Base de Dados (Último Upload)")

        if st.button("Limpar Base de Dados"):
            arquivos_para_remover = [
                "saldo.csv",
                "perfil.csv",
                "ordens.csv",
                "previsao.csv"
            ]

            removidos = []

            for arq in arquivos_para_remover:
                if os.path.exists(arq):
                    os.remove(arq)
                    removidos.append(arq)

            if removidos:
                st.success(f"Arquivos removidos: {', '.join(removidos)}")
            else:
                st.warning("Nenhum arquivo encontrado para remover.")

        arquivos = {
            "Saldo": ("saldo.csv", "Codigo"),
            "Perfil": ("perfil.csv", "Item"),
            "Ordens": ("ordens.csv", None),
            "Previsão": ("previsao.csv", "COD")
        }

        for nome, (arquivo, chave) in arquivos.items():
            st.subheader(nome)

            if os.path.exists(arquivo):
                df = pd.read_csv(arquivo)

                if "Data Processamento" in df.columns:
                    df = df.sort_values(
                        by=["Data Processamento", "Hora Processamento"],
                        ascending=False
                    )

                if chave and chave in df.columns:
                    df = df.drop_duplicates(subset=[chave], keep="first")

                st.dataframe(df, use_container_width=True)
                botao_downloads(df, f"{nome}_limpo", nome)
            else:
                st.warning(f"{nome} ainda não carregado.")

    st.markdown("---")
    if st.button("Ir para Dashboard PCP", use_container_width=True):
        ir_para("dashboard")


def render_dashboard():
    st.title("Dashboard PCP")
    st.caption("Página com endereço próprio")

    try:
        saldo = pd.read_csv("saldo.csv")
        perfil = pd.read_csv("perfil.csv")
        previsao = pd.read_csv("previsao.csv")
    except Exception:
        st.warning("Faça upload dos dados primeiro na página principal.")
        if st.button("Voltar para Página Principal", use_container_width=True):
            ir_para("principal")
        st.stop()

    saldo = saldo.sort_values(
        by=["Data Processamento", "Hora Processamento"],
        ascending=False
    ).drop_duplicates(subset=["Codigo"])

    perfil = perfil.sort_values(
        by=["Data Processamento", "Hora Processamento"],
        ascending=False
    )

    previsao = previsao.sort_values(
        by=["Data Processamento", "Hora Processamento"],
        ascending=False
    ).drop_duplicates(subset=["COD"])

    base = previsao[["COD", "PRODUTO"]].copy()
    base.columns = ["Codigo", "Descricao"]

    saldo_base = saldo[["Codigo", "Saldo Total", "Saldo Almox 3"]]

    perfil["Quantidade"] = (
        perfil["Quantidade"]
        .astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
        .astype(float)
    )

    perfil["Data Fim"] = pd.to_datetime(perfil["Data Fim"], dayfirst=True)

    perfil["Referencia"] = (
        perfil["Data Fim"].dt.isocalendar().week.astype(str).str.zfill(2)
        + "."
        + perfil["Data Fim"].dt.year.astype(str)
    )

    semana = agora().isocalendar()[1]
    ano = agora().year
    ref = str(semana).zfill(2) + "." + str(ano)

    dc = perfil[perfil["Tipo"] == "DC"].groupby("Item")["Quantidade"].sum().reset_index()
    dc.columns = ["Codigo", "Demanda Pedido"]

    dp = perfil[perfil["Tipo"] == "DP"].groupby("Item")["Quantidade"].sum().reset_index()
    dp.columns = ["Codigo", "Demanda DP"]

    dp_sem = perfil[
        (perfil["Tipo"] == "DP") &
        (perfil["Referencia"] == ref)
    ].groupby("Item")["Quantidade"].sum().reset_index()
    dp_sem.columns = ["Codigo", "DP Semana Atual"]

    df = base.merge(saldo_base, on="Codigo", how="left")
    df = df.merge(dc, on="Codigo", how="left")
    df = df.merge(dp, on="Codigo", how="left")
    df = df.merge(dp_sem, on="Codigo", how="left")

    df = df.fillna(0)
    df["Saldo vs Demanda"] = df["Saldo Almox 3"] - df["Demanda Pedido"]

    def status(row):
        if row["Saldo vs Demanda"] < 0:
            return "FALTA"
        if row["Demanda Pedido"] >= row["Saldo Almox 3"] * 0.5:
            return "RISCO"
        return "OK"

    df["Status"] = df.apply(status, axis=1)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total de Itens", len(df))
    col2.metric("Itens em Falta", int((df["Status"] == "FALTA").sum()))
    col3.metric("Itens em Risco", int((df["Status"] == "RISCO").sum()))
    col4.metric("Itens OK", int((df["Status"] == "OK").sum()))

    st.markdown("---")

    opcoes_status = ["FALTA", "RISCO", "OK"]
    opcoes_disponiveis = [item for item in opcoes_status if item in df["Status"].unique()]

    status_selecionado = st.multiselect(
        "Filtrar Status",
        options=opcoes_disponiveis,
        default=opcoes_disponiveis
    )

    texto_busca = st.text_input("Buscar por código ou descrição")

    if status_selecionado:
        df_filtrado = df[df["Status"].isin(status_selecionado)].copy()
    else:
        df_filtrado = df.iloc[0:0].copy()

    if texto_busca:
        filtro = texto_busca.strip().lower()
        df_filtrado = df_filtrado[
            df_filtrado["Codigo"].astype(str).str.lower().str.contains(filtro, na=False) |
            df_filtrado["Descricao"].astype(str).str.lower().str.contains(filtro, na=False)
        ]

    st.bar_chart(df["Status"].value_counts())

    st.dataframe(df_filtrado, use_container_width=True)
    botao_downloads(df_filtrado, "analise_pcp_filtrada", "Analise_PCP")

    st.markdown("---")
    if st.button("Voltar para Página Principal", use_container_width=True):
        ir_para("principal")


menu_lateral()

if pagina_atual() == "dashboard":
    render_dashboard()
else:
    render_principal()

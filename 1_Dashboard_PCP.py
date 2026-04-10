from datetime import datetime
from io import BytesIO
import re

import pandas as pd
import pytz
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="Dashboard PCP", layout="wide")

fuso = pytz.timezone("America/Sao_Paulo")


def agora():
    return datetime.now(fuso)


def exportar_excel_formatado(df, nome_aba="Dados"):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nome_aba)
        ws = writer.book[nome_aba]

        ws.freeze_panes = "A2"

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if ws.max_row >= 2 and ws.max_column >= 1:
            ultima_coluna = get_column_letter(ws.max_column)
            nome_tabela = re.sub(r"\W+", "", nome_aba) or "Dados"
            tabela = Table(
                displayName=f"Tabela{nome_tabela[:20]}",
                ref=f"A1:{ultima_coluna}{ws.max_row}"
            )
            tabela.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(tabela)

    output.seek(0)
    return output.getvalue()


def botao_downloads(df, nome_base, nome_aba):
    col_csv, col_excel = st.columns(2)

    with col_csv:
        st.download_button(
            f"Baixar {nome_base} CSV",
            df.to_csv(index=False).encode("utf-8"),
            file_name=f"{nome_base}.csv",
            mime="text/csv",
            key=f"csv_{nome_base}_{nome_aba}"
        )

    with col_excel:
        st.download_button(
            f"Baixar {nome_base} Excel",
            exportar_excel_formatado(df, nome_aba=nome_aba[:31]),
            file_name=f"{nome_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"excel_{nome_base}_{nome_aba}"
        )


st.title("Dashboard PCP")

try:
    saldo = pd.read_csv("saldo.csv")
    perfil = pd.read_csv("perfil.csv")
    previsao = pd.read_csv("previsao.csv")
except Exception:
    st.warning("Faça upload dos dados primeiro na página principal.")
    st.stop()

saldo = saldo.sort_values(
    by=["Data Processamento", "Hora Processamento"],
    ascending=False
).drop_duplicates(subset=["Codigo"])

perfil = perfil.sort_values(
    by=["Data Processamento", "Hora Processamento"],
    ascending=False
)

previsao = previsao.sort_values(
    by=["Data Processamento", "Hora Processamento"],
    ascending=False
).drop_duplicates(subset=["COD"])

base = previsao[["COD", "PRODUTO"]].copy()
base.columns = ["Codigo", "Descricao"]

saldo_base = saldo[["Codigo", "Saldo Total", "Saldo Almox 3"]]

perfil["Quantidade"] = (
    perfil["Quantidade"]
    .astype(str)
    .str.replace(".", "", regex=False)
    .str.replace(",", ".", regex=False)
    .astype(float)
)

perfil["Data Fim"] = pd.to_datetime(perfil["Data Fim"], dayfirst=True)

perfil["Referencia"] = (
    perfil["Data Fim"].dt.isocalendar().week.astype(str).str.zfill(2)
    + "."
    + perfil["Data Fim"].dt.year.astype(str)
)

semana = agora().isocalendar()[1]
ano = agora().year
ref = str(semana).zfill(2) + "." + str(ano)

dc = perfil[perfil["Tipo"] == "DC"].groupby("Item")["Quantidade"].sum().reset_index()
dc.columns = ["Codigo", "Demanda Pedido"]

dp = perfil[perfil["Tipo"] == "DP"].groupby("Item")["Quantidade"].sum().reset_index()
dp.columns = ["Codigo", "Demanda DP"]

dp_sem = perfil[
    (perfil["Tipo"] == "DP") &
    (perfil["Referencia"] == ref)
].groupby("Item")["Quantidade"].sum().reset_index()
dp_sem.columns = ["Codigo", "DP Semana Atual"]

df = base.merge(saldo_base, on="Codigo", how="left")
df = df.merge(dc, on="Codigo", how="left")
df = df.merge(dp, on="Codigo", how="left")
df = df.merge(dp_sem, on="Codigo", how="left")

df = df.fillna(0)
df["Saldo vs Demanda"] = df["Saldo Almox 3"] - df["Demanda Pedido"]


def status(row):
    if row["Saldo vs Demanda"] < 0:
        return "FALTA"
    if row["Demanda Pedido"] >= row["Saldo Almox 3"] * 0.5:
        return "RISCO"
    return "OK"


df["Status"] = df.apply(status, axis=1)

col1, col2, col3, col4 = st.columns(4)
col1.metric("Total de Itens", len(df))
col2.metric("Itens em Falta", int((df["Status"] == "FALTA").sum()))
col3.metric("Itens em Risco", int((df["Status"] == "RISCO").sum()))
col4.metric("Itens OK", int((df["Status"] == "OK").sum()))

st.markdown("---")

opcoes_status = ["FALTA", "RISCO", "OK"]
opcoes_disponiveis = [item for item in opcoes_status if item in df["Status"].unique()]

status_selecionado = st.multiselect(
    "Filtrar Status",
    options=opcoes_disponiveis,
    default=opcoes_disponiveis
)

texto_busca = st.text_input("Buscar por código ou descrição")

if status_selecionado:
    df_filtrado = df[df["Status"].isin(status_selecionado)].copy()
else:
    df_filtrado = df.iloc[0:0].copy()

if texto_busca:
    filtro = texto_busca.strip().lower()
    df_filtrado = df_filtrado[
        df_filtrado["Codigo"].astype(str).str.lower().str.contains(filtro, na=False) |
        df_filtrado["Descricao"].astype(str).str.lower().str.contains(filtro, na=False)
    ]

st.bar_chart(df["Status"].value_counts())
st.dataframe(df_filtrado, use_container_width=True)

botao_downloads(df_filtrado, "dashboard_pcp", "Dashboard_PCP")

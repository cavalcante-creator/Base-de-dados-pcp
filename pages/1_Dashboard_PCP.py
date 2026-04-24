import os
import re
from datetime import datetime
from io import BytesIO, StringIO

import pandas as pd
import pdfplumber
import pytz
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="PCP Produção", layout="wide")

st.markdown("""
<style>
    .main {
        background: linear-gradient(180deg, #f7fafc 0%, #eef4f7 100%);
    }

    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }

    h1, h2, h3 {
        color: #12344d;
        font-weight: 700;
    }

    div[data-testid="stAlert"] {
        border-radius: 14px;
        border: 1px solid #d7e3ee;
    }

    div.stButton > button {
        background: linear-gradient(90deg, #0f766e, #0ea5a4);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.6rem 1.2rem;
        font-weight: 600;
    }

    div.stButton > button:hover {
        background: linear-gradient(90deg, #0b5f59, #0b8f8d);
        color: white;
    }

    div[data-testid="stDownloadButton"] > button {
        background: white;
        color: #12344d;
        border: 1px solid #c9d9e6;
        border-radius: 12px;
        font-weight: 600;
    }

    div[data-testid="stFileUploader"] {
        background: white;
        border: 1px dashed #aac4d6;
        border-radius: 16px;
        padding: 10px;
    }

    div[data-testid="stDataFrame"] {
        background: white;
        border-radius: 16px;
        padding: 8px;
        border: 1px solid #dbe7f0;
        box-shadow: 0 4px 14px rgba(18, 52, 77, 0.05);
    }

    .custom-card {
        background: white;
        padding: 20px;
        border-radius: 18px;
        border: 1px solid #dbe7f0;
        box-shadow: 0 6px 20px rgba(18, 52, 77, 0.08);
        margin-bottom: 16px;
    }

    .small-title {
        font-size: 14px;
        color: #5b7488;
        margin-bottom: 6px;
    }

    .big-number {
        font-size: 32px;
        font-weight: 700;
        color: #12344d;
    }
</style>
""", unsafe_allow_html=True)

fuso = pytz.timezone("America/Sao_Paulo")

def tratar_numero(valor):
    if pd.isna(valor):
        return 0
    valor = str(valor).strip()
    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except:
        return 0

def agora():
    return datetime.now(fuso)


def exportar_excel_formatado(df, nome_aba="Dados"):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nome_aba)
        ws = writer.book[nome_aba]

        ws.freeze_panes = "A2"

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if ws.max_row >= 2 and ws.max_column >= 1:
            ultima_coluna = get_column_letter(ws.max_column)
            nome_tabela = re.sub(r"\W+", "", nome_aba) or "Dados"
            tabela = Table(
                displayName=f"Tabela{nome_tabela[:20]}",
                ref=f"A1:{ultima_coluna}{ws.max_row}"
            )
            tabela.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(tabela)

        for idx, coluna in enumerate(df.columns, start=1):
            letra_coluna = get_column_letter(idx)
            valores = [str(coluna)]
            valores.extend("" if pd.isna(valor) else str(valor) for valor in df[coluna])
            ws.column_dimensions[letra_coluna].width = min(max(len(valor) for valor in valores) + 2, 40)

            if pd.api.types.is_numeric_dtype(df[coluna]):
                for cell in ws[letra_coluna][1:]:
                    cell.number_format = "#,##0.00"

        if "Status" in df.columns:
            idx_status = list(df.columns).index("Status") + 1
            for cell in ws[get_column_letter(idx_status)][1:]:
                valor = str(cell.value or "")
                if "FALTA" in valor:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
                elif "RISCO" in valor:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
                elif "OK" in valor:
                    cell.fill = PatternFill(fill_type="solid", fgColor="C6E0B4")

    output.seek(0)
    return output.getvalue()


def botao_downloads(df, nome_base, nome_aba):
    col_csv, col_excel = st.columns(2)

    with col_csv:
        st.download_button(
            f"Baixar {nome_base} CSV",
            df.to_csv(index=False).encode("utf-8"),
            file_name=f"{nome_base}.csv",
            mime="text/csv",
            key=f"csv_{nome_base}_{nome_aba}"
        )

    with col_excel:
        st.download_button(
            f"Baixar {nome_base} Excel",
            exportar_excel_formatado(df, nome_aba=nome_aba[:31]),
            file_name=f"{nome_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"excel_{nome_base}_{nome_aba}"
        )


st.markdown("""
<div class="custom-card">
    <div class="small-title">Sistema PCP</div>
    <div class="big-number">PCP Produção</div>
    <div style="color:#5b7488; margin-top:8px;">
        Faça os uploads, processe os arquivos e acompanhe a base consolidada.
    </div>
</div>
""", unsafe_allow_html=True)

st.info("No menu lateral do Streamlit haverá uma página separada chamada Dashboard PCP.")

abas = st.tabs([
    "Saldo",
    "Perfil",
    "Ordens",
    "Previsão",
    "Parâmetros",
    "Base de Dados"
])

with abas[0]:
    st.title("Saldo Produção")

    file = st.file_uploader("PDF Saldo", type=["pdf"])

    if file:
        with open("saldo_temp.pdf", "wb") as f:
            f.write(file.read())

        if st.button("Processar Saldo"):
            linhas = []

            with pdfplumber.open("saldo_temp.pdf") as pdf:
                for p in pdf.pages:
                    texto = p.extract_text()
                    if texto:
                        linhas.extend(texto.split("\n"))

            dados = {}
            codigo_atual = None

            for linha in linhas:
                linha = linha.strip()
                if not linha:
                    continue

                if "ALMOXARIFADO" in linha.upper():
                    codigo_match = None
                else:
                    codigo_match = re.search(r"\b([A-Z]{1,3}\d{3,5})\b", linha)

                if codigo_match:
                    codigo_atual = codigo_match.group(1)
                    if codigo_atual not in dados:
                        dados[codigo_atual] = {
                            "Codigo": codigo_atual,
                            "Saldo Total": 0,
                            "Saldo Almox 3": 0,
                            "Saldo Almox 30": 0
                        }
                    continue

                if "ALMOXARIFADO" in linha.upper() and codigo_atual:
                    nums = re.findall(r"[\d\.]+\,\d+", linha)
                    if nums:
                        valor = float(nums[-1].replace(".", "").replace(",", "."))
                        dados[codigo_atual]["Saldo Total"] += valor

                        if re.search(r"ALMOXARIFADO\s*:\s*30\b", linha.upper()):
                            dados[codigo_atual]["Saldo Almox 30"] += valor

                        if re.search(r"ALMOXARIFADO\s*:\s*3\b", linha.upper()):
                            dados[codigo_atual]["Saldo Almox 3"] += valor

            df = pd.DataFrame(dados.values())
            df["Data Processamento"] = agora().strftime("%d/%m/%Y")
            df["Hora Processamento"] = agora().strftime("%H:%M:%S")

            df.to_csv("saldo.csv", index=False)

            st.success("Saldo processado com sucesso.")
            st.dataframe(df, use_container_width=True)

with abas[1]:
    st.title("Perfil Produção")

    file = st.file_uploader("PDF Perfil", type=["pdf"])

    if file:
        with open("perfil_temp.pdf", "wb") as f:
            f.write(file.read())

        if st.button("Processar Perfil"):
            movimentacoes = []
            codigo_item = ""

            regex = re.compile(
                r"(DD|DC|DP).*?(\d{2}/\d{2}/\d{4}).*?(-?[\d,.]+)"
            )

            with pdfplumber.open("perfil_temp.pdf") as pdf:
                for p in pdf.pages:
                    texto = p.extract_text()
                    if texto:
                        for linha in texto.split("\n"):
                            item_match = re.search(r"Item:\s*(\S+)", linha)
                            if item_match:
                                codigo_item = item_match.group(1)

                            mov = regex.search(linha)
                            if mov:
                                movimentacoes.append({
                                    "Item": codigo_item,
                                    "Tipo": mov.group(1),
                                    "Data Fim": mov.group(2),
                                    "Quantidade": mov.group(3)
                                })

            df = pd.DataFrame(movimentacoes)
            df["Data Processamento"] = agora().strftime("%d/%m/%Y")
            df["Hora Processamento"] = agora().strftime("%H:%M:%S")

            df.to_csv("perfil.csv", index=False)

            st.success("Perfil processado com sucesso.")
            st.dataframe(df, use_container_width=True)

with abas[2]:
    st.title("Ordens")

    file = st.file_uploader("CSV Ordens", type=["csv"])

    if file:
        conteudo = file.read().decode("utf-8", errors="ignore")
        df = pd.read_csv(StringIO(conteudo), sep=None, engine="python")

        df["Data Processamento"] = agora().strftime("%d/%m/%Y")
        df["Hora Processamento"] = agora().strftime("%H:%M:%S")

        df.to_csv("ordens.csv", index=False)

        st.success("Ordens carregadas com sucesso.")
        st.dataframe(df, use_container_width=True)

with abas[3]:
    st.title("Previsão")

    file = st.file_uploader("Excel Previsão", type=["xlsx"])

    if file:
        df_raw = pd.read_excel(file, header=None)

        linha_header = None
        for i in range(len(df_raw)):
            if "COD" in df_raw.iloc[i].astype(str).str.upper().values:
                linha_header = i
                break

        if linha_header is None:
            st.error("Não foi possível localizar a linha de cabeçalho.")
            st.stop()

        df = pd.read_excel(file, header=linha_header)
        df.columns = df.columns.astype(str).str.upper().str.strip()

        col_cod = [c for c in df.columns if "COD" in c][0]
        col_prod = [c for c in df.columns if "PROD" in c][0]

        df = df[[col_cod, col_prod]]
        df.columns = ["COD", "PRODUTO"]

        df["Data Processamento"] = agora().strftime("%d/%m/%Y")
        df["Hora Processamento"] = agora().strftime("%H:%M:%S")

        df.to_csv("previsao.csv", index=False)

        st.success("Previsão carregada com sucesso.")
        st.dataframe(df, use_container_width=True)

with abas[4]:
    st.title("Parâmetros")

    file = st.file_uploader("Excel/XLS Parâmetros", type=["xls", "xlsx"])

    if file:
        conteudo = file.read()

        try:
            texto_html = conteudo.decode("iso-8859-1")
            eh_html = "<html" in texto_html.lower() or "<table" in texto_html.lower()
        except Exception:
            eh_html = False

        if eh_html:
            df = pd.read_html(BytesIO(conteudo), encoding="iso-8859-1", header=0)[0]
            df.columns = df.columns.astype(str).str.strip()

            colunas_num = ["LOTE MIN", "LOTE MAX", "LOTE MULT", "ESTQ SEG", "TEMP REP", "TEMP SEG", "AGRUP", "CONS MEDIO"]
            for col in colunas_num:
                if col in df.columns:
                    df[col] = df[col].apply(tratar_numero)
        else:
            df = pd.read_excel(BytesIO(conteudo))

        df["Data Processamento"] = agora().strftime("%d/%m/%Y")
        df["Hora Processamento"] = agora().strftime("%H:%M:%S")

        df.to_csv("parametros.csv", index=False)

        st.success("Parâmetros carregados com sucesso.")
        st.dataframe(df, use_container_width=True)

with abas[5]:
    st.title("Base de Dados (Último Upload)")

    if st.button("Limpar Base de Dados"):
        arquivos_para_remover = [
            "saldo.csv",
            "perfil.csv",
            "ordens.csv",
            "previsao.csv",
            "parametros.csv"
        ]

        removidos = []

        for arq in arquivos_para_remover:
            if os.path.exists(arq):
                os.remove(arq)
                removidos.append(arq)

        if removidos:
            st.success(f"Arquivos removidos: {', '.join(removidos)}")
        else:
            st.warning("Nenhum arquivo encontrado para remover.")

    arquivos = {
        "Saldo": ("saldo.csv", "Codigo"),
        "Perfil": ("perfil.csv", "Item"),
        "Ordens": ("ordens.csv", None),
        "Previsão": ("previsao.csv", "COD"),
        "Parâmetros": ("parametros.csv", "COD ITEM")
    }

    for nome, (arquivo, chave) in arquivos.items():
        st.subheader(nome)

        if os.path.exists(arquivo):
            df = pd.read_csv(arquivo)

            if "Data Processamento" in df.columns:
                df = df.sort_values(
                    by=["Data Processamento", "Hora Processamento"],
                    ascending=False
                )

            if chave and chave in df.columns:
                df = df.drop_duplicates(subset=[chave], keep="first")

            st.dataframe(df, use_container_width=True)
            botao_downloads(df, f"{nome}_limpo", nome)
        else:
            st.warning(f"{nome} ainda não carregado.")
